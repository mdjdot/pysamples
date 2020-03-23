from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

"""
https://openpyxl.readthedocs.io/en/stable/#
"""


def writeExcel():
    wb = Workbook()
    ws = wb.active

    ws.append(["name", "Age"])
    ws["A2"] = "张三"
    ws["B2"] = "18"
    ws.append(["李四", 22])

    ws["D1"] = datetime.datetime.now()

    wb.save("./sample.xlsx")


def readExcel():
    wb = load_workbook("./sample.xlsx")
    ws = wb._sheets[0]
    
    print(ws["A1"].value)


if __name__ == "__main__":
    writeExcel()
    readExcel()
